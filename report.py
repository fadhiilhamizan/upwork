"""Excel report generation with openpyxl.

Computed values are written directly rather than as live formulas, so the file
is correct without a spreadsheet engine ever opening it.
"""

from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

import config
from models import Job

FONT_NAME = "Arial"

UPWORK_GREEN = "14A800"
HEADER_FILL = PatternFill("solid", fgColor=UPWORK_GREEN)
TITLE_COLOR = "0D5C00"
SUBTLE = "595959"
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")

FIT_STYLES = {
    "High": (PatternFill("solid", fgColor="C6EFCE"), "006100"),
    "Medium": (PatternFill("solid", fgColor="FFEB9C"), "9C6500"),
    "Low": (PatternFill("solid", fgColor="FFC7CE"), "9C0006"),
}

NEW_FILL = PatternFill("solid", fgColor="DDEBF7")

THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

JOB_COLUMNS = [
    ("Score", 8),
    ("Fit", 10),
    ("New", 7),
    ("Title", 62),
    ("Type", 10),
    ("Rate / Budget", 20),
    ("Proposals", 14),
    ("Posted", 18),
    ("Location", 18),
    ("Skills", 40),
    ("Description", 70),
]


def _write(ws, row, col, value, *, bold=False, size=10, color="000000",
           fill=None, align=None, wrap=False, border=False, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=FONT_NAME, bold=bold, size=size, color=color)
    if fill is not None:
        cell.fill = fill
    cell.alignment = Alignment(
        horizontal=align, vertical="top" if wrap else "center", wrap_text=wrap
    )
    if border:
        cell.border = CELL_BORDER
    if number_format:
        cell.number_format = number_format
    return cell


def _hyperlink(cell, url: str) -> None:
    if not url:
        return
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url)
    cell.font = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")


def _header_row(ws, row, headers) -> None:
    for index, (label, width) in enumerate(headers, start=1):
        _write(ws, row, index, label, bold=True, color="FFFFFF",
               fill=HEADER_FILL, align="center", border=True)
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[row].height = 22


def _write_job_rows(ws, jobs: List[Job], start_row: int) -> None:
    for offset, job in enumerate(jobs):
        row = start_row + offset
        fill, font_color = FIT_STYLES.get(job.fit, (None, "000000"))

        _write(ws, row, 1, job.score, align="center", border=True,
               number_format="0.0")
        _write(ws, row, 2, job.fit, bold=True, color=font_color, fill=fill,
               align="center", border=True)
        _write(ws, row, 3, "NEW" if job.is_new else "", bold=True,
               color="1F4E79", fill=NEW_FILL if job.is_new else None,
               align="center", border=True)

        title_cell = _write(ws, row, 4, job.title, wrap=True, border=True)
        _hyperlink(title_cell, job.url)

        _write(ws, row, 5, job.price_type, align="center", border=True)
        _write(ws, row, 6, job.rate, border=True)
        _write(ws, row, 7, job.proposals_text or "n/a", align="center", border=True)
        _write(ws, row, 8, job.posted_text or "n/a", border=True)
        _write(ws, row, 9, job.location or "n/a", border=True)
        _write(ws, row, 10, job.skills_text, wrap=True, border=True)
        _write(ws, row, 11, job.description, wrap=True, border=True)

        ws.row_dimensions[row].height = 46


def _build_category_sheet(wb: Workbook, category: str, jobs: List[Job]) -> None:
    ws = wb.create_sheet(title=category[:31])

    _write(ws, 1, 1, config.CATEGORY_DISPLAY.get(category, category),
           bold=True, size=14, color=TITLE_COLOR)
    _write(ws, 2, 1, f"{len(jobs)} job(s), sorted by score", size=9, color=SUBTLE)

    _header_row(ws, 4, JOB_COLUMNS)

    if jobs:
        _write_job_rows(ws, jobs, 5)
        ws.auto_filter.ref = f"A4:{get_column_letter(len(JOB_COLUMNS))}{4 + len(jobs)}"
    else:
        _write(ws, 5, 1, "No jobs matched this category in the last "
                         f"{config.MAX_JOB_AGE_DAYS} days.", color=SUBTLE)

    ws.freeze_panes = "D5"


def _build_dashboard(wb: Workbook, by_category: Dict[str, List[Job]],
                     all_jobs: List[Job], generated_at: datetime) -> None:
    ws = wb.create_sheet(title="Dashboard", index=0)
    ws.sheet_view.showGridLines = False

    _write(ws, 1, 1, "Upwork Daily Job Report", bold=True, size=18,
           color=TITLE_COLOR)
    _write(ws, 2, 1, f"Generated: {generated_at.strftime('%A, %d %B %Y at %H:%M:%S')}",
           size=10, color=SUBTLE)
    _write(ws, 3, 1, f"Filters: {config.FILTER_SUMMARY}", size=9, color=SUBTLE)

    new_count = sum(1 for job in all_jobs if job.is_new)
    _write(ws, 4, 1,
           f"{len(all_jobs)} job(s) after dedupe, {new_count} new since the last run",
           size=10, bold=True)

    for width, letter in zip([26, 10, 10, 12, 10, 10, 14, 12, 16],
                             "ABCDEFGHI"):
        ws.column_dimensions[letter].width = width

    # -- per category counts ------------------------------------------------
    table_top = 6
    _write(ws, table_top, 1, "Jobs by Category", bold=True, size=12,
           color=TITLE_COLOR)

    head_row = table_top + 1
    for index, label in enumerate(["Category", "Jobs", "High", "Medium", "Low", "New"],
                                  start=1):
        _write(ws, head_row, index, label, bold=True, color="FFFFFF",
               fill=HEADER_FILL, align="center", border=True)

    for offset, category in enumerate(config.CATEGORY_ORDER):
        row = head_row + 1 + offset
        jobs = by_category.get(category, [])
        band = BAND_FILL if offset % 2 else None
        _write(ws, row, 1, config.CATEGORY_DISPLAY.get(category, category),
               fill=band, border=True)
        _write(ws, row, 2, len(jobs), fill=band, align="center", border=True)
        for column, fit in [(3, "High"), (4, "Medium"), (5, "Low")]:
            _write(ws, row, column, sum(1 for j in jobs if j.fit == fit),
                   fill=band, align="center", border=True)
        _write(ws, row, 6, sum(1 for j in jobs if j.is_new), fill=band,
               align="center", border=True)

    total_row = head_row + 1 + len(config.CATEGORY_ORDER)
    _write(ws, total_row, 1, "TOTAL", bold=True, border=True)
    _write(ws, total_row, 2, len(all_jobs), bold=True, align="center", border=True)
    for column, fit in [(3, "High"), (4, "Medium"), (5, "Low")]:
        _write(ws, total_row, column,
               sum(1 for j in all_jobs if j.fit == fit),
               bold=True, align="center", border=True)
    _write(ws, total_row, 6, new_count, bold=True, align="center", border=True)

    # -- bar chart ----------------------------------------------------------
    chart = BarChart()
    chart.type = "col"
    chart.title = "Jobs per Category"
    chart.y_axis.title = "Jobs"
    chart.x_axis.title = None
    chart.height = 8
    chart.width = 18
    chart.legend = None

    data = Reference(ws, min_col=2, min_row=head_row,
                     max_row=total_row - 1)
    categories = Reference(ws, min_col=1, min_row=head_row + 1,
                           max_row=total_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, f"H{table_top}")

    # -- top 10 -------------------------------------------------------------
    top_top = total_row + 3
    _write(ws, top_top, 1, "Top 10 Picks", bold=True, size=12, color=TITLE_COLOR)

    top_head = top_top + 1
    headers = ["#", "Score", "Fit", "Title", "Category", "Proposals", "Posted",
               "Rate / Budget"]
    for index, label in enumerate(headers, start=1):
        _write(ws, top_head, index, label, bold=True, color="FFFFFF",
               fill=HEADER_FILL, align="center", border=True)

    top_jobs = sorted(all_jobs, key=lambda j: j.score, reverse=True)[:10]
    if not top_jobs:
        _write(ws, top_head + 1, 1, "No jobs found on this run.", color=SUBTLE)
    for offset, job in enumerate(top_jobs):
        row = top_head + 1 + offset
        fill, font_color = FIT_STYLES.get(job.fit, (None, "000000"))
        _write(ws, row, 1, offset + 1, align="center", border=True)
        _write(ws, row, 2, job.score, align="center", border=True,
               number_format="0.0")
        _write(ws, row, 3, job.fit, bold=True, color=font_color, fill=fill,
               align="center", border=True)
        title_cell = _write(ws, row, 4, job.title, wrap=True, border=True)
        _hyperlink(title_cell, job.url)
        _write(ws, row, 5, job.category, wrap=True, border=True)
        _write(ws, row, 6, job.proposals_text or "n/a", align="center", border=True)
        _write(ws, row, 7, job.posted_text or "n/a", border=True)
        _write(ws, row, 8, job.rate, border=True)
        ws.row_dimensions[row].height = 30

    # The title column needs the room on this sheet.
    ws.column_dimensions["D"].width = 58


def build_workbook(jobs: List[Job], output_path=None,
                   generated_at: datetime = None) -> str:
    """Write every job into the single report workbook, overwriting it."""
    output_path = output_path or config.OUTPUT_PATH
    generated_at = generated_at or datetime.now()

    by_category: Dict[str, List[Job]] = {c: [] for c in config.CATEGORY_ORDER}
    for job in jobs:
        by_category.setdefault(job.category, []).append(job)
    for category in by_category:
        by_category[category].sort(key=lambda j: j.score, reverse=True)

    wb = Workbook()
    wb.remove(wb.active)

    _build_dashboard(wb, by_category, jobs, generated_at)
    for category in config.CATEGORY_ORDER:
        _build_category_sheet(wb, category, by_category.get(category, []))

    try:
        wb.save(str(output_path))
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write {output_path}. It is probably open in Excel, "
            "close it and run the script again."
        ) from exc

    return str(output_path)
